"""
CSV to Formatted Excel (.xlsx) Converter

Converts CSV files to professionally formatted Excel workbooks with:
- Auto-detected delimiters
- Auto-width columns
- Header formatting (bold, background color)
- Data type detection (numbers, dates, text)
- Frozen header row
- Auto-filters
- Optional chart generation from numeric columns

Requirements: pip install openpyxl

Usage:
    python csv-to-xlsx.py input.csv
    python csv-to-xlsx.py input.csv -o output.xlsx
    python csv-to-xlsx.py input.csv --header-color 2E74B5 --header-font-color FFFFFF
    python csv-to-xlsx.py input.csv --chart --chart-type bar
    python csv-to-xlsx.py input.tsv --delimiter "\t"
"""

import argparse
import csv
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DATE_PATTERNS = [
    (r"^\d{4}-\d{2}-\d{2}$", "%Y-%m-%d"),
    (r"^\d{2}/\d{2}/\d{4}$", "%m/%d/%Y"),
    (r"^\d{2}-\d{2}-\d{4}$", "%m-%d-%Y"),
    (r"^\d{4}/\d{2}/\d{2}$", "%Y/%m/%d"),
    (r"^\d{2}\.\d{2}\.\d{4}$", "%d.%m.%Y"),
    (r"^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}$", "%Y-%m-%dT%H:%M:%S"),
    (r"^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}$", "%Y-%m-%d %H:%M:%S"),
]


def detect_delimiter(file_path: str) -> str:
    with open(file_path, "r", encoding="utf-8-sig") as f:
        sample = f.read(8192)

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        return dialect.delimiter
    except csv.Error:
        return ","


def parse_value(value: str):
    if not value or value.strip() == "":
        return None

    stripped = value.strip()

    for pattern, fmt in DATE_PATTERNS:
        if re.match(pattern, stripped):
            try:
                return datetime.strptime(stripped, fmt)
            except ValueError:
                pass

    # Percentage: "85.3%" → 0.853
    if stripped.endswith("%"):
        try:
            return float(stripped[:-1]) / 100.0
        except ValueError:
            pass

    # Currency-prefixed numbers: "$1,234.56" → 1234.56
    if stripped and stripped[0] in "$€£¥":
        try:
            return float(stripped[1:].replace(",", "").strip())
        except ValueError:
            pass

    # Plain numbers (with optional commas as thousands separator)
    try:
        cleaned = stripped.replace(",", "")
        if "." in cleaned or cleaned.lstrip("-").isdigit():
            num = float(cleaned)
            if num == int(num) and "." not in stripped:
                return int(num)
            return num
    except ValueError:
        pass

    return stripped


def auto_width(ws, min_width: int = 8, max_width: int = 50, padding: int = 3):
    for col_cells in ws.columns:
        lengths = []
        for cell in col_cells:
            if cell.value is not None:
                cell_len = len(str(cell.value))
                lengths.append(cell_len)
        if lengths:
            optimal = min(max(max(lengths) + padding, min_width), max_width)
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = optimal


def apply_header_style(ws, row: int, bg_color: str, font_color: str):
    header_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    header_font = Font(bold=True, color=font_color, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="999999"),
        right=Side(style="thin", color="DDDDDD"),
    )

    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border


def apply_data_formatting(ws, start_row: int):
    light_gray = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    data_alignment = Alignment(vertical="center")
    thin_border = Border(
        bottom=Side(style="thin", color="EEEEEE"),
    )

    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row), start=0):
        for cell in row:
            cell.alignment = data_alignment
            cell.border = thin_border

            if isinstance(cell.value, datetime):
                cell.number_format = "YYYY-MM-DD"
            elif isinstance(cell.value, float) and 0 <= cell.value <= 1:
                # Heuristic: small floats from percentage parsing get % format
                raw_str = str(cell.value)
                if len(raw_str.split(".")[-1]) <= 4:
                    cell.number_format = "0.0%"
            elif isinstance(cell.value, float):
                cell.number_format = "#,##0.00"
            elif isinstance(cell.value, int) and abs(cell.value) >= 1000:
                cell.number_format = "#,##0"

        if row_idx % 2 == 1:
            for cell in row:
                if cell.fill == PatternFill():
                    cell.fill = light_gray


def find_numeric_columns(ws, header_row: int, data_start_row: int):
    numeric_cols = []
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(row=header_row, column=col_idx).value
        if not header:
            continue

        sample_count = 0
        numeric_count = 0
        for row_idx in range(data_start_row, min(data_start_row + 20, ws.max_row + 1)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                sample_count += 1
                if isinstance(val, (int, float)):
                    numeric_count += 1

        if sample_count > 0 and numeric_count / sample_count >= 0.8:
            numeric_cols.append((col_idx, header))

    return numeric_cols


def add_chart(ws, chart_type: str, numeric_cols, data_start_row: int, max_row: int):
    if not numeric_cols or max_row <= data_start_row:
        return

    # Limit to first 5 numeric columns to keep chart readable
    cols_to_chart = numeric_cols[:5]

    chart_classes = {
        "bar": BarChart,
        "line": LineChart,
        "pie": PieChart,
    }
    ChartClass = chart_classes.get(chart_type, BarChart)
    chart = ChartClass()
    chart.title = "Data Summary"
    chart.width = 20
    chart.height = 12

    if chart_type == "pie" and cols_to_chart:
        col_idx, col_name = cols_to_chart[0]
        data_ref = Reference(ws, min_col=col_idx, min_row=data_start_row - 1, max_row=max_row)
        cat_ref = Reference(ws, min_col=1, min_row=data_start_row, max_row=max_row)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cat_ref)
        chart.title = f"{col_name} Distribution"
    else:
        for col_idx, col_name in cols_to_chart:
            data_ref = Reference(ws, min_col=col_idx, min_row=data_start_row - 1, max_row=max_row)
            chart.add_data(data_ref, titles_from_data=True)

        cat_ref = Reference(ws, min_col=1, min_row=data_start_row, max_row=max_row)
        chart.set_categories(cat_ref)

        if hasattr(chart, "x_axis"):
            chart.x_axis.title = str(ws.cell(row=1, column=1).value or "Category")
        if hasattr(chart, "y_axis"):
            chart.y_axis.title = "Value"

    chart_anchor = f"A{max_row + 3}"
    ws.add_chart(chart, chart_anchor)


def convert_csv_to_xlsx(
    input_path: str,
    output_path: str,
    delimiter: str | None = None,
    header_color: str = "2E74B5",
    header_font_color: str = "FFFFFF",
    sheet_name: str = "Data",
    generate_chart: bool = False,
    chart_type: str = "bar",
):
    input_file = Path(input_path)
    if not input_file.exists():
        print(f"Error: Input file '{input_path}' not found.", file=sys.stderr)
        sys.exit(1)

    if delimiter is None:
        delimiter = detect_delimiter(input_path)

    if delimiter == "\\t":
        delimiter = "\t"

    print(f"Detected delimiter: {repr(delimiter)}")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    with open(input_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f, delimiter=delimiter)
        for row_idx, row in enumerate(reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                if row_idx == 1:
                    ws.cell(row=row_idx, column=col_idx, value=value.strip())
                else:
                    ws.cell(row=row_idx, column=col_idx, value=parse_value(value))

    if ws.max_row < 1:
        print("Warning: CSV file appears to be empty.", file=sys.stderr)
        wb.save(output_path)
        return

    apply_header_style(ws, row=1, bg_color=header_color, font_color=header_font_color)

    if ws.max_row > 1:
        apply_data_formatting(ws, start_row=2)

    ws.freeze_panes = "A2"

    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    auto_width(ws)

    ws.sheet_properties.tabColor = header_color

    if generate_chart:
        numeric_cols = find_numeric_columns(ws, header_row=1, data_start_row=2)
        if numeric_cols:
            add_chart(ws, chart_type, numeric_cols, data_start_row=2, max_row=ws.max_row)
            print(f"Chart added with {len(numeric_cols[:5])} numeric column(s).")
        else:
            print("No numeric columns detected for chart generation.")

    wb.save(output_path)
    row_count = ws.max_row - 1
    col_count = ws.max_column
    print(f"Converted {row_count} rows x {col_count} columns → {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Convert CSV to formatted Excel (.xlsx)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python csv-to-xlsx.py sales.csv
  python csv-to-xlsx.py data.tsv -o report.xlsx --delimiter "\\t"
  python csv-to-xlsx.py metrics.csv --chart --chart-type line
  python csv-to-xlsx.py report.csv --header-color 4472C4 --header-font-color FFFFFF
        """,
    )
    parser.add_argument("input", help="Path to input CSV file")
    parser.add_argument("-o", "--output", help="Path to output .xlsx file (default: same name as input with .xlsx extension)")
    parser.add_argument("-d", "--delimiter", default=None, help="CSV delimiter (auto-detected if omitted). Use '\\t' for tab.")
    parser.add_argument("--header-color", default="2E74B5", help="Header background color as hex (default: 2E74B5)")
    parser.add_argument("--header-font-color", default="FFFFFF", help="Header font color as hex (default: FFFFFF)")
    parser.add_argument("--sheet-name", default="Data", help="Worksheet name (default: Data)")
    parser.add_argument("--chart", action="store_true", help="Generate a chart from numeric columns")
    parser.add_argument("--chart-type", choices=["bar", "line", "pie"], default="bar", help="Chart type (default: bar)")

    args = parser.parse_args()

    output = args.output or str(Path(args.input).with_suffix(".xlsx"))

    convert_csv_to_xlsx(
        input_path=args.input,
        output_path=output,
        delimiter=args.delimiter,
        header_color=args.header_color,
        header_font_color=args.header_font_color,
        sheet_name=args.sheet_name,
        generate_chart=args.chart,
        chart_type=args.chart_type,
    )


if __name__ == "__main__":
    main()
